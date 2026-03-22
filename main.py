"""Gmail Reader — Claude анализирует входящие письма и расставляет теги."""

import asyncio
import json
import re
import sys

from claude_agent_sdk import (
    AssistantMessage,
    ClaudeAgentOptions,
    ClaudeSDKClient,
    ResultMessage,
    TextBlock,
)

import gmail_client

# Маппинг системных тегов → человекочитаемые названия
SYSTEM_LABEL_NAMES = {
    "INBOX": "Входящие",
    "SENT": "Отправленные",
    "DRAFT": "Черновики",
    "TRASH": "Корзина",
    "SPAM": "Спам",
    "STARRED": "Помеченные",
    "IMPORTANT": "Важные",
    "UNREAD": "Непрочитанные",
    "CATEGORY_PERSONAL": "Личные",
    "CATEGORY_SOCIAL": "Соцсети",
    "CATEGORY_PROMOTIONS": "Промоакции",
    "CATEGORY_UPDATES": "Обновления",
    "CATEGORY_FORUMS": "Форумы",
}

SYSTEM_PROMPT = """Ты — профессиональный ассистент, который отличает рассылку от не-рассылки.

Для каждого письма определи категорию:
- 🟢 ЛИЧНОЕ — письмо адресовано лично, требует ответа или внимания
- 🔵 ПОЛЕЗНАЯ РАССЫЛКА — интересный контент, на который подписан осознанно
- 🟡 УВЕДОМЛЕНИЕ (важное) — транзакционное письмо, которое стоит прочитать (упоминания, security alerts, важные PR review, billing)
- 🟠 УВЕДОМЛЕНИЕ (шум) — однотипные повторяющиеся алерты, CI/CD крашит контейнеры, массовые GitHub пуши, уведомления о деплоях — полезно в теории, но засоряет inbox
- 🔴 МУСОР — маркетинговые рассылки, промоакции, спам

Для каждого письма укажи:
1. Категорию (одну из пяти выше)
2. Краткое резюме (1 предложение)
3. Рекомендацию: оставить / отписаться / настроить фильтр

В конце дай сводку: сколько писем в каждой категории и общую рекомендацию по наведению порядка.

Отвечай на русском языке."""

SYSTEM_PROMPT_WITH_TAGS = """Ты — профессиональный ассистент, который отличает рассылку от не-рассылки.

Для каждого письма определи категорию:
- 🟢 ЛИЧНОЕ — письмо адресовано лично, требует ответа или внимания
- 🔵 ПОЛЕЗНАЯ РАССЫЛКА — интересный контент, на который подписан осознанно
- 🟡 УВЕДОМЛЕНИЕ (важное) — транзакционное письмо, которое стоит прочитать (упоминания, security alerts, важные PR review, billing)
- 🟠 УВЕДОМЛЕНИЕ (шум) — однотипные повторяющиеся алерты, CI/CD крашит контейнеры, массовые GitHub пуши, уведомления о деплоях — полезно в теории, но засоряет inbox
- 🔴 МУСОР — маркетинговые рассылки, промоакции, спам

Для каждого письма укажи:
1. Категорию (одну из пяти выше)
2. Краткое резюме (1 предложение)
3. Рекомендацию: оставить / отписаться / настроить фильтр
4. Подходящие пользовательские теги из доступного списка

В конце дай сводку: сколько писем в каждой категории и общую рекомендацию по наведению порядка.

ВАЖНО: После текстового анализа обязательно добавь JSON-блок с рекомендациями по тегам в формате:
```json
{"tagging": [{"message_id": "ID_ПИСЬМА", "add_labels": ["Label_ID1", "Label_ID2"]}]}
```
Указывай только те теги, которых ещё нет у письма. Используй ID тегов (не имена).
Если письму не нужны новые теги, не включай его в список.

Отвечай на русском языке."""


SYSTEM_PROMPT_SENDERS = """Ты — аналитик email-рассылок. Тебе дан список отправителей с примерами тем их писем.

Для каждого отправителя определи:
1. Тематику (краткое описание, 2-5 слов)
2. Сектор (одна из категорий ниже или предложи свою, если ни одна не подходит)

Стандартные секторы:
- Банки и финансы
- Криптовалюта и трейдинг
- Хостинг и облака
- DevOps и инфраструктура
- SaaS и инструменты
- Маркетплейсы и e-commerce
- Соцсети
- Образование
- HR и рекрутинг
- Путешествия и бронирование
- Медиа и контент
- AI и ML
- Безопасность
- Госуслуги и налоги
- Другое

Верни результат СТРОГО в JSON формате:
```json
{"senders": [{"email": "адрес@example.com", "theme": "тематика", "sector": "Сектор"}]}
```

ВАЖНО: включи ВСЕХ отправителей из списка, не пропуская ни одного. Отвечай на русском языке."""


def format_senders_prompt(senders_subjects, counts):
    """Форматирует список отправителей с темами для анализа Claude."""
    lines = [f"Проанализируй {len(senders_subjects)} отправителей рассылок:\n"]
    for email, subjects in senders_subjects.items():
        count = counts.get(email, 0)
        lines.append(f"--- {email} ({count} писем) ---")
        lines.append(f"Темы: {', '.join(repr(s) for s in subjects[:10])}")
        lines.append("")
    return "\n".join(lines)


def parse_senders_response(text):
    """Извлечь JSON с анализом отправителей из ответа Claude."""
    pattern = r'```json\s*(\{[^`]*"senders"[^`]*\})\s*```'
    match = re.search(pattern, text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1))
        except json.JSONDecodeError:
            pass
    pattern2 = r'(\{"senders"\s*:\s*\[.*?\]\s*\})'
    match2 = re.search(pattern2, text, re.DOTALL)
    if match2:
        try:
            return json.loads(match2.group(1))
        except json.JSONDecodeError:
            pass
    return None


def export_senders_to_excel(analysis, counts, filename="senders_analysis.xlsx"):
    """Экспорт анализа отправителей в Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ отправителей"

    # Заголовки
    headers = ["№", "Отправитель", "Кол-во писем", "Тематика", "Сектор"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Сортировка: по сектору, внутри — по кол-ву (desc)
    senders = analysis.get("senders", [])
    for s in senders:
        s["count"] = counts.get(s.get("email", ""), 0)
    senders.sort(key=lambda x: (x.get("sector", "Другое"), -x.get("count", 0)))

    # Данные
    current_sector = None
    row = 2
    num = 1
    sector_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for s in senders:
        sector = s.get("sector", "Другое")
        if sector != current_sector:
            # Строка-разделитель с названием сектора
            cell = ws.cell(row=row, column=1, value=sector)
            cell.font = Font(bold=True, size=11)
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = sector_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1
            current_sector = sector

        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=s.get("email", ""))
        ws.cell(row=row, column=3, value=s.get("count", 0))
        ws.cell(row=row, column=4, value=s.get("theme", ""))
        ws.cell(row=row, column=5, value=sector)
        num += 1
        row += 1

    # Ширина колонок
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 28

    wb.save(filename)
    return filename


def _label_id_to_name(label_id, labels_map):
    """Преобразовать ID тега в человекочитаемое название."""
    if label_id in SYSTEM_LABEL_NAMES:
        return SYSTEM_LABEL_NAMES[label_id]
    return labels_map.get(label_id, label_id)


def format_emails_prompt(emails, all_labels):
    """Форматирует письма в промпт для Claude (с информацией о тегах)."""
    # Собираем маппинг id → name для всех тегов
    labels_map = {lb["id"]: lb["name"] for lb in all_labels}

    # Список пользовательских тегов
    user_labels = [lb for lb in all_labels if lb["type"] == "user"]

    lines = [f"Проанализируй {len(emails)} входящих писем:\n"]

    # Доступные пользовательские теги
    lines.append("Доступные пользовательские теги для назначения:")
    for lb in user_labels:
        lines.append(f"  - {lb['name']} (ID: {lb['id']})")
    lines.append("")

    for i, email in enumerate(emails, 1):
        # Текущие теги письма
        current_labels = [
            _label_id_to_name(lid, labels_map) for lid in email.get("labelIds", [])
        ]

        lines.append(f"--- Письмо {i} ---")
        lines.append(f"ID: {email['id']}")
        lines.append(f"От: {email['from']}")
        lines.append(f"Дата: {email['date']}")
        lines.append(f"Тема: {email['subject']}")
        lines.append(f"Текущие теги: {', '.join(current_labels) or 'нет'}")
        lines.append(f"Текст:\n{email['body']}")
        lines.append("")
    return "\n".join(lines)


def format_emails_prompt_simple(emails):
    """Форматирует письма в промпт для Claude (только анализ, без тегов)."""
    lines = [f"Проанализируй {len(emails)} входящих писем:\n"]

    for i, email in enumerate(emails, 1):
        lines.append(f"--- Письмо {i} ---")
        lines.append(f"От: {email['from']}")
        lines.append(f"Дата: {email['date']}")
        lines.append(f"Тема: {email['subject']}")
        lines.append(f"Текст:\n{email['body']}")
        lines.append("")
    return "\n".join(lines)


def parse_tagging_response(text):
    """Извлечь JSON-блок с тегами из ответа Claude."""
    # Ищем JSON в блоке ```json ... ``` или просто {"tagging": ...}
    pattern = r'```json\s*(\{[^`]*"tagging"[^`]*\})\s*```'
    match = re.search(pattern, text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1))
        except json.JSONDecodeError:
            pass

    # Fallback: ищем JSON без обёртки в код-блок
    pattern2 = r'(\{"tagging"\s*:\s*\[.*?\]\s*\})'
    match2 = re.search(pattern2, text, re.DOTALL)
    if match2:
        try:
            return json.loads(match2.group(1))
        except json.JSONDecodeError:
            pass

    return None


def apply_tags(tagging_data, labels_map):
    """Применить теги к письмам после подтверждения пользователя."""
    actions = tagging_data.get("tagging", [])
    if not actions:
        print("\nНет рекомендаций по тегам.")
        return

    print("\n--- Рекомендации по тегам ---")
    for action in actions:
        msg_id = action.get("message_id", "???")
        add = action.get("add_labels", [])
        label_names = [labels_map.get(lid, lid) for lid in add]
        if label_names:
            print(f"  Письмо {msg_id[:12]}...: + {', '.join(label_names)}")

    answer = input("\nПрименить теги? (y/n): ").strip().lower()
    if answer not in ("y", "д", "yes", "да"):
        print("Отменено.")
        return

    for action in actions:
        msg_id = action.get("message_id")
        add = action.get("add_labels", [])
        if msg_id and add:
            try:
                gmail_client.modify_message_labels(msg_id, add_label_ids=add)
                print(f"  Теги применены к {msg_id[:12]}...")
            except Exception as e:
                print(f"  Ошибка для {msg_id[:12]}...: {e}")

    print("Готово!")


# ── CLI: управление тегами ──


def cmd_labels_list():
    """Показать все теги."""
    labels = gmail_client.get_labels()
    system_labels = [lb for lb in labels if lb["type"] == "system"]
    user_labels = [lb for lb in labels if lb["type"] == "user"]

    print("=== Системные теги ===")
    for lb in sorted(system_labels, key=lambda x: x["name"]):
        readable = SYSTEM_LABEL_NAMES.get(lb["id"], "")
        suffix = f" ({readable})" if readable else ""
        print(f"  {lb['id']:<30} {lb['name']}{suffix}")

    print(f"\n=== Пользовательские теги ({len(user_labels)}) ===")
    for lb in sorted(user_labels, key=lambda x: x["name"]):
        print(f"  {lb['id']:<30} {lb['name']}")


def cmd_labels_create(name):
    """Создать тег."""
    result = gmail_client.create_label(name)
    print(f"Тег создан: {result['name']} (ID: {result['id']})")


def cmd_labels_delete(label_id):
    """Удалить тег."""
    gmail_client.delete_label(label_id)
    print(f"Тег {label_id} удалён.")


# ── Основной flow ──


async def cmd_analyze_senders(top_n):
    """Анализ отправителей рассылок: тематика, сектор, экспорт в Excel."""
    print("Поиск самых активных отправителей рассылок...")
    try:
        counter = gmail_client.get_top_senders()
    except RuntimeError as e:
        print(f"Ошибка: {e}")
        return

    if not counter:
        print("Писем не найдено.")
        return

    top = counter.most_common(top_n)
    senders_list = [email for email, _ in top]
    print(f"Топ-{len(top)} отправителей. Загрузка тем писем...")

    try:
        subjects = gmail_client.get_subjects_by_senders(senders_list)
    except RuntimeError as e:
        print(f"Ошибка: {e}")
        return

    if not subjects:
        print("Не удалось загрузить темы писем.")
        return

    print(f"Темы загружены для {len(subjects)} отправителей. Отправляю Claude на анализ...\n")

    prompt = format_senders_prompt(subjects, counter)

    options = ClaudeAgentOptions(
        system_prompt=SYSTEM_PROMPT_SENDERS,
        allowed_tools=[],
        permission_mode="default",
    )

    full_response = ""
    async with ClaudeSDKClient(options) as client:
        await client.query(prompt)
        async for message in client.receive_response():
            if isinstance(message, AssistantMessage):
                for block in message.content:
                    if isinstance(block, TextBlock):
                        print(block.text, end="", flush=True)
                        full_response += block.text
            elif isinstance(message, ResultMessage):
                if message.total_cost_usd is not None:
                    print(f"\n\n[Стоимость запроса: ${message.total_cost_usd:.4f}]")

    analysis = parse_senders_response(full_response)
    if not analysis:
        print("\nНе удалось распарсить JSON из ответа Claude.")
        return

    filename = export_senders_to_excel(analysis, counter)
    print(f"\nРезультаты сохранены в {filename}")


def print_help():
    """Вывести справку по использованию."""
    print("""Gmail Reader — Claude анализирует входящие письма и расставляет теги.

Использование:
  python main.py                              анализ писем (INBOX)
  python main.py --tag                        анализ + авто-тегирование
  python main.py --tag --label CATEGORY_X     авто-тегирование писем с указанным тегом
  python main.py --label CATEGORY_UPDATES     анализ писем с указанным тегом

  python main.py top [N]                      топ-N отправителей рассылок (по умолчанию 10)
  python main.py analyze-senders [N]          анализ отправителей по секторам → Excel (по умолчанию 50)
  python main.py mark <Label_ID> <emails...>  пометить письма от отправителей тегом
  python main.py mark-query <Label_ID> <query> пометить письма по Gmail search query

  python main.py labels                       список всех тегов
  python main.py labels create <Имя>          создать пользовательский тег
  python main.py labels delete <Label_ID>     удалить тег

Флаги:
  --tag          включить авто-тегирование (Claude предложит теги, вы подтвердите)
  --label <ID>   фильтровать письма по тегу
  --help, -h     показать эту справку""")


async def main():
    args = sys.argv[1:]

    # --help / -h
    if "--help" in args or "-h" in args:
        print_help()
        return

    # CLI: top-senders subcommand
    if args and args[0] == "top":
        top_n = 10
        if len(args) >= 2:
            try:
                top_n = int(args[1])
            except ValueError:
                print(f"Ошибка: '{args[1]}' — не число. Использование: python main.py top [N]")
                return
        print("Поиск самых активных отправителей рассылок...")
        try:
            counter = gmail_client.get_top_senders()
        except RuntimeError as e:
            print(f"Ошибка: {e}")
            return
        if not counter:
            print("Писем не найдено.")
            return
        top = counter.most_common(top_n)
        print(f"\n{'№':>3}  {'Кол-во':>6}  Отправитель")
        print(f"{'—'*3}  {'—'*6}  {'—'*40}")
        for i, (sender, count) in enumerate(top, 1):
            print(f"{i:>3}  {count:>6}  {sender}")
        total = sum(counter.values())
        top_total = sum(count for _, count in top)
        print(f"\nИтого: {total} писем от {len(counter)} отправителей")
        print(f"Топ-{len(top)} покрывают {top_total} писем ({top_total*100//total}%)")
        return

    # CLI: mark subcommand
    if args and args[0] == "mark":
        if len(args) < 3:
            print("Использование: python main.py mark <Label_ID> <email1> [email2] ...")
            print("Пример: python main.py mark Label_12 alert@uptimerobot.com noreply@booking.com")
            return
        label_id = args[1]
        senders = args[2:]
        print(f"Помечаю письма от {len(senders)} отправителей тегом {label_id}...")
        try:
            results = gmail_client.label_messages_from_senders(label_id, senders)
        except Exception as e:
            print(f"Ошибка: {e}")
            return
        total = 0
        for sender, count in results.items():
            print(f"  {sender}: {count} писем")
            total += count
        print(f"\nИтого помечено: {total} писем")
        return

    # CLI: mark-query subcommand
    if args and args[0] == "mark-query":
        if len(args) < 3:
            print("Использование: python main.py mark-query <Label_ID> <gmail_query>")
            print('Пример: python main.py mark-query Label_12 "from:notifications@github.com subject:[axisrow/"')
            return
        label_id = args[1]
        query = " ".join(args[2:])
        print(f"Запрос: {query}")
        print(f"Помечаю тегом {label_id}...")
        try:
            count = gmail_client.label_messages_by_query(label_id, query)
        except Exception as e:
            print(f"Ошибка: {e}")
            return
        print(f"Помечено: {count} писем")
        return

    # CLI: analyze-senders subcommand
    if args and args[0] == "analyze-senders":
        top_n = 50
        if len(args) >= 2:
            try:
                top_n = int(args[1])
            except ValueError:
                print(f"Ошибка: '{args[1]}' — не число. Использование: python main.py analyze-senders [N]")
                return
        await cmd_analyze_senders(top_n)
        return

    # CLI: labels subcommand
    if args and args[0] == "labels":
        if len(args) == 1:
            cmd_labels_list()
        elif args[1] == "create" and len(args) >= 3:
            cmd_labels_create(" ".join(args[2:]))
        elif args[1] == "delete" and len(args) >= 3:
            cmd_labels_delete(args[2])
        else:
            print("Использование:")
            print("  python main.py labels              — список тегов")
            print("  python main.py labels create Имя   — создать тег")
            print("  python main.py labels delete ID    — удалить тег")
        return

    # Парсим флаги
    do_tag = "--tag" in args
    label_ids = None
    if "--label" in args:
        idx = args.index("--label")
        if idx + 1 < len(args):
            label_ids = [args[idx + 1]]
        else:
            print("Ошибка: --label требует аргумент (ID тега)")
            return

    print("Загрузка писем из Gmail...")
    try:
        emails = gmail_client.get_emails(max_results=10, label_ids=label_ids)
    except RuntimeError as e:
        print(f"Ошибка: {e}")
        return

    if not emails:
        print("Входящих писем не найдено.")
        return

    print(f"Загружено {len(emails)} писем.")

    if do_tag:
        # Загружаем теги для тегирования
        print("Загрузка тегов...")
        try:
            all_labels = gmail_client.get_labels()
        except Exception as e:
            print(f"Ошибка загрузки тегов: {e}")
            all_labels = []
        labels_map = {lb["id"]: lb["name"] for lb in all_labels}
        prompt = format_emails_prompt(emails, all_labels)
        system_prompt = SYSTEM_PROMPT_WITH_TAGS
    else:
        prompt = format_emails_prompt_simple(emails)
        system_prompt = SYSTEM_PROMPT

    print("Отправляю Claude на анализ...\n")

    options = ClaudeAgentOptions(
        system_prompt=system_prompt,
        allowed_tools=[],
        permission_mode="default",
    )

    full_response = ""

    async with ClaudeSDKClient(options) as client:
        await client.query(prompt)

        async for message in client.receive_response():
            if isinstance(message, AssistantMessage):
                for block in message.content:
                    if isinstance(block, TextBlock):
                        print(block.text, end="", flush=True)
                        full_response += block.text
            elif isinstance(message, ResultMessage):
                if message.total_cost_usd is not None:
                    print(f"\n\n[Стоимость запроса: ${message.total_cost_usd:.4f}]")

    # Парсим и применяем теги (только в режиме --tag)
    if do_tag:
        tagging_data = parse_tagging_response(full_response)
        if tagging_data:
            apply_tags(tagging_data, labels_map)
        else:
            print("\n(Claude не предложил изменений тегов)")


if __name__ == "__main__":
    asyncio.run(main())
